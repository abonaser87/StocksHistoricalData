import sqlite3
import openpyxl
import datetime
from umalqurra.hijri_date import HijriDate

wb = openpyxl.load_workbook('FR.xlsx',data_only=True)
conn = sqlite3.connect('Test.db')
c = conn.cursor()

def create_table():
    c.execute("""CREATE TABLE IF NOT EXISTS company (
    id      integer PRIMARY KEY,
    name    varchar(255)
)""")
    c.execute("""CREATE TABLE IF NOT EXISTS statement (
    id      integer PRIMARY KEY,
    name    varchar(255)
)""")
    c.execute("""CREATE TABLE IF NOT EXISTS statementRow (
    id      integer  PRIMARY KEY,
    statementId integer ,
    rowOrder  integer ,
    rowTitle  varchar(255) ,
    rowDescription varchar(255) NULL,
    rowProperties varchar(255) NULL,
    FOREIGN KEY (statementId) REFERENCES statement (id)
)""")

    c.execute("""CREATE TABLE IF NOT EXISTS statementFact (
    id integer PRIMARY KEY,
    companyId      integer ,
    statementRowId integer ,
    entrydate         date ,
    amount         numeric ,
    FOREIGN KEY (companyId) REFERENCES company (id),
    FOREIGN KEY (statementRowId) REFERENCES statementRow (id)
)""")


create_table()


for sheet in wb.get_sheet_names():
    wSheet = wb.get_sheet_by_name(sheet)
    if sheet.startswith("!"):
        continue
    cols = wSheet.columns
    companyId = str(wSheet.title)
    try:
        c.execute("INSERT INTO company (name) VALUES(?)", (companyId,))
        c.execute("SELECT id FROM company WHERE name=?", (companyId,))
        companyId = c.fetchone()[0]
    except:
        c.execute("SELECT id FROM company WHERE name=?", (companyId,))
        companyId = c.fetchone()[0]
        pass

    for col in cols:
        i = 1
        j = 1
        loop = True
        for cell in col:
            if cell.value == "BALANCE SHEET":
                xy = openpyxl.utils.coordinate_from_string(cell.coordinate)
                row = xy[1]
                col = openpyxl.utils.column_index_from_string(xy[0])
                while loop == True:
                    try:
                        c.execute("INSERT INTO statementRow (statementId,rowOrder,rowTitle) SELECT 1," + str(
                            i) + ",\"" + wSheet.cell(row=row + i,
                                                     column=col).value + "\" WHERE NOT EXISTS(SELECT 1 FROM statementRow where statementId=1 AND rowOrder=" + str(
                            i) + " AND rowTitle=\"" + wSheet.cell(row=row + i, column=col).value + "\")")
                    except:
                        pass
                    j = 1
                    while wSheet.cell(row=row, column=col + j).value != None:
                        amount = wSheet.cell(row=row + i, column=col + j).value
                        if amount == "-":
                            amount = 0
                        entrydate = wSheet.cell(row=row, column=col + j).value
                        entrydate = entrydate.strftime("%Y-%m-%d")
                        c.execute("SELECT id FROM statementRow WHERE rowTitle=? AND statementId=1",
                                  (wSheet.cell(row=row + i, column=col).value,))
                        sRowId = c.fetchone()[0]
                        try:
                            c.execute("INSERT INTO statementFact (companyId,statementRowId,entrydate,amount) VALUES (?,?,?,?)",(companyId,sRowId,entrydate,amount))
                            conn.commit()
                        except sqlite3.Error as e:
                            print e.message
                            pass
                        j = j + 1
                    i = i + 1
                    if wSheet.cell(row=row + i, column=col).value == None:
                        loop = False

            if cell.value == "STATEMENT OF INCOME":
                xy = openpyxl.utils.coordinate_from_string(cell.coordinate)
                row = xy[1]
                col = openpyxl.utils.column_index_from_string(xy[0])
                while loop == True:
                    try:
                        c.execute("INSERT INTO statementRow (statementId,rowOrder,rowTitle) SELECT 2," + str(
                            i) + ",\"" + wSheet.cell(row=row + i,
                                                     column=col).value + "\" WHERE NOT EXISTS(SELECT 1 FROM statementRow where statementId=2 AND rowOrder=" + str(
                            i) + " AND rowTitle=\"" + wSheet.cell(row=row + i, column=col).value + "\")")
                    except:
                        pass
                    j = 1
                    while wSheet.cell(row=row, column=col + j).value != None:
                        amount = wSheet.cell(row=row + i, column=col + j).value
                        if amount == "-":
                            amount = 0
                        entrydate = wSheet.cell(row=row, column=col + j).value
                        entrydate = entrydate.strftime("%Y-%m-%d")
                        c.execute("SELECT id FROM statementRow WHERE rowTitle=? AND statementId=2",
                                  (wSheet.cell(row=row + i, column=col).value,))
                        sRowId = c.fetchone()[0]
                        try:
                            c.execute("INSERT INTO statementFact (companyId,statementRowId,entrydate,amount) VALUES (?,?,?,?)",(companyId,sRowId,entrydate,amount))
                            conn.commit()
                        except sqlite3.Error as e:
                            print e.message
                            pass
                        j = j + 1
                    i = i + 1
                    if wSheet.cell(row=row + i, column=col).value == None:
                        loop = False

            if cell.value == "CASH FLOW":
                xy = openpyxl.utils.coordinate_from_string(cell.coordinate)
                row = xy[1]
                col = openpyxl.utils.column_index_from_string(xy[0])
                while loop == True:
                    try:
                        c.execute("INSERT INTO statementRow (statementId,rowOrder,rowTitle) SELECT 3," + str(
                            i) + ",\"" + wSheet.cell(row=row + i,
                                                     column=col).value + "\" WHERE NOT EXISTS(SELECT 1 FROM statementRow where statementId=3 AND rowOrder=" + str(
                            i) + " AND rowTitle=\"" + wSheet.cell(row=row + i, column=col).value + "\")")
                    except:
                        pass
                    j = 1
                    while wSheet.cell(row=row, column=col + j).value != None:
                        amount = wSheet.cell(row=row + i, column=col + j).value
                        if amount == "-":
                            amount = 0
                        entrydate = wSheet.cell(row=row, column=col + j).value
                        entrydate = entrydate.strftime("%Y-%m-%d")
                        c.execute("SELECT id FROM statementRow WHERE rowTitle=? AND statementId=3",
                                  (wSheet.cell(row=row + i, column=col).value,))
                        sRowId = c.fetchone()[0]
                        try:
                            c.execute("INSERT INTO statementFact (companyId,statementRowId,entrydate,amount) VALUES (?,?,?,?)",(companyId,sRowId,entrydate,amount))
                            conn.commit()
                        except sqlite3.Error as e:
                            print e.message
                            pass
                        j = j + 1
                    i = i + 1
                    if wSheet.cell(row=row + i, column=col).value == None or wSheet.cell(row=row + i - 1,
                                                                                         column=col).value == "Cash at End of Period":
                        loop = False
conn.commit()
c.close()
conn.close()
# Dynamically add data from excel sheet
